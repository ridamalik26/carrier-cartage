"""Scored output workbook (Section 8) — reproduces the column shape of
Working_of_performance_evaluation.xlsx -> Sheet1."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.scoring.scoring_engine import ContractorScore

COLUMNS = [
    "Carrier", "Carriage", "Total loads", "kms travelled", "Fleet", "OGRA fleet",
    "Accident", "Abnormal shortage", "Fake documents", "seal temp",
    "Quality Fail & Unauthorized Mod.", "OMC Loading", "HSE Observations", "ATS Trained",
    "Accident %", "DDT %", "Medical Screening %", "OGRA fleet %", "HSE Observation",
    "Quality Fail & Unauthorized Mod %", "Abnormal Shortage2", "Fake Documents2",
    "Seal Tempering", "Other OMC Loading", "Overall Score", "Previous Score",
    "Status (Flagged/Cleared)",
]

FLAGGED_FILL = PatternFill(start_color="FFC1443C", end_color="FFC1443C", fill_type="solid")
CLEARED_FILL = PatternFill(start_color="FF4FA8A0", end_color="FF4FA8A0", fill_type="solid")
PERCENT_COLUMNS = {
    "Accident %", "DDT %", "Medical Screening %", "OGRA fleet %", "HSE Observation",
    "Quality Fail & Unauthorized Mod %", "Abnormal Shortage2", "Fake Documents2",
    "Seal Tempering", "Other OMC Loading", "Overall Score", "Previous Score",
}


def build_workbook(
    contractor_rows: list[dict],
    scores: dict[int, ContractorScore],
    previous_scores: dict[int, float] | None = None,
) -> Workbook:
    previous_scores = previous_scores or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    col_idx = {name: i + 1 for i, name in enumerate(COLUMNS)}

    for row in contractor_rows:
        cc = row["cc_number"]
        score = scores[cc]
        s = score.category_scores
        prev = previous_scores.get(cc)

        ws.append([
            cc,
            row.get("name", str(cc)),
            row.get("total_loads", 0),
            row.get("kms_travelled", 0),
            row.get("fleet", 0),
            row.get("ogra_fleet", 0),
            row.get("accident_count", 0),
            row.get("abnormal_shortage_count", 0),
            row.get("fake_documents_count", 0),
            row.get("seal_temp_count", 0),
            row.get("quality_fail_count", 0),
            row.get("omc_count", 0),
            row.get("hse_count", 0),
            row.get("ats_trained", 0),
            s["accident"],
            s["ddt"],
            s["medical_screening"],
            s["ogra_fleet"],
            s["hse"],
            s["quality_fail"],
            s["abnormal_shortage"],
            s["fake_documents"],
            s["seal_temp"],
            s["other_omc"],
            score.overall_score,
            prev if prev is not None else "",
            "Flagged" if score.flagged else "Cleared",
        ])

        excel_row = ws.max_row
        status_cell = ws.cell(row=excel_row, column=col_idx["Status (Flagged/Cleared)"])
        status_cell.fill = FLAGGED_FILL if score.flagged else CLEARED_FILL
        status_cell.font = Font(color="FFFFFFFF", bold=True)

        for percent_col in PERCENT_COLUMNS:
            cell = ws.cell(row=excel_row, column=col_idx[percent_col])
            if isinstance(cell.value, (int, float)) and cell.value != "":
                cell.number_format = "0.00%"

    for i, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(name) + 2)

    return wb


def save_workbook(contractor_rows: list[dict], scores: dict[int, ContractorScore], output_path: str,
                   previous_scores: dict[int, float] | None = None) -> str:
    wb = build_workbook(contractor_rows, scores, previous_scores)
    wb.save(output_path)
    return output_path
