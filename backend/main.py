"""FastAPI app wiring the ingestion -> scoring -> reporting -> email pipeline
(Section 11/12). Upload raw files, calculate scores, download the scored
workbook, and drive the review-and-send queue for flagged contractors.

Two storage modes, chosen automatically at import time:
- Local disk (default): everything lives in data/ and the in-memory `state`
  singleton persists for as long as the process runs (start.bat/start.sh).
- Vercel (when BLOB_READ_WRITE_TOKEN + KV_REST_API_URL/TOKEN are set): a
  serverless invocation can't rely on local disk or in-memory state surviving
  to the next request, so `state` is round-tripped through Vercel KV on every
  /api/* call (see auth_gate below) and workbooks are stored in Vercel Blob
  instead of data/output.
"""
from __future__ import annotations

import io
import json
import os
import shutil
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook
from fastapi import FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from backend import blob_store, kv_store
from backend.auth import issue_token, revoke_token, verify_credentials, verify_token
from backend.ingestion.abnormal_shortage import ingest_abnormal_shortage
from backend.ingestion.fake_documents import ingest_fake_documents
from backend.ingestion.master_data import ingest_master_data
from backend.ingestion.quality_fail import ingest_quality_fail
from backend.ingestion.seal_tempering import ingest_seal_tempering
from backend.email_service.graph_mailer import send_report_email
from backend.reporting.docx_report import generate_report
from backend.reporting.excel_output import build_workbook
from backend.scoring.formula_config import default_formula_config
from backend.scoring.scoring_engine import score_all_contractors
from backend.state import MANUAL_CATEGORIES, state

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_DIR = os.path.join(BASE_DIR, "data", "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "output")
REPORTS_DIR = os.path.join(BASE_DIR, "data", "output", "reports")
HISTORY_DIR = os.path.join(BASE_DIR, "data", "output", "history")
HISTORY_MANIFEST_PATH = os.path.join(HISTORY_DIR, "manifest.json")
CONTRACTOR_EMAILS_PATH = os.path.join(BASE_DIR, "data", "contractor_emails.json")

load_dotenv(os.path.join(BASE_DIR, ".env"))

# Presence of these env vars is what decides local-disk vs. Vercel storage —
# see module docstring. Set by adding the Vercel KV + Blob integrations to
# the project (Storage tab in the Vercel dashboard).
REMOTE_STORAGE = bool(os.environ.get("BLOB_READ_WRITE_TOKEN")) and bool(os.environ.get("KV_REST_API_URL"))
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"

if not REMOTE_STORAGE:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(REPORTS_DIR, exist_ok=True)
    os.makedirs(HISTORY_DIR, exist_ok=True)

# Excel columns (see backend/reporting/excel_output.py COLUMNS) that hold each
# category's score, in the same order calculate_contractor_score built them —
# two of them (Abnormal Shortage2 / Fake Documents2) got a "2" suffix in the
# sheet to avoid clashing with the raw count columns of the same name.
HISTORY_CATEGORY_COLUMNS = [
    ("Accident %", "accident"),
    ("DDT %", "ddt"),
    ("Medical Screening %", "medical_screening"),
    ("OGRA fleet %", "ogra_fleet"),
    ("HSE Observation", "hse"),
    ("Quality Fail & Unauthorized Mod %", "quality_fail"),
    ("Abnormal Shortage2", "abnormal_shortage"),
    ("Fake Documents2", "fake_documents"),
    ("Seal Tempering", "seal_temp"),
    ("Other OMC Loading", "other_omc"),
]


def load_history_manifest() -> dict:
    if REMOTE_STORAGE:
        return kv_store.get_json("history_manifest", {})
    if not os.path.exists(HISTORY_MANIFEST_PATH):
        return {}
    with open(HISTORY_MANIFEST_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_history_manifest(manifest: dict) -> None:
    if REMOTE_STORAGE:
        kv_store.set_json("history_manifest", manifest)
        return
    with open(HISTORY_MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)


def _parse_history_workbook(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col_idx = {name: i for i, name in enumerate(header)}

    results = []
    for row in rows_iter:
        if row[col_idx["Carrier"]] is None:
            continue
        category_scores = {
            key: round((row[col_idx[col_name]] or 0) * 100, 2)
            for col_name, key in HISTORY_CATEGORY_COLUMNS if col_name in col_idx
        }
        results.append({
            "cc_number": row[col_idx["Carrier"]],
            "name": row[col_idx["Carriage"]],
            "overall_score": round((row[col_idx["Overall Score"]] or 0) * 100, 2),
            "category_scores": category_scores,
        })
    wb.close()
    return results


def load_contractor_emails() -> None:
    if REMOTE_STORAGE:
        state.contractor_emails = {
            int(cc): email for cc, email in kv_store.get_json("contractor_emails", {}).items()
        }
        return
    if not os.path.exists(CONTRACTOR_EMAILS_PATH):
        return
    with open(CONTRACTOR_EMAILS_PATH, "r", encoding="utf-8") as f:
        raw = json.load(f)
    state.contractor_emails = {int(cc): email for cc, email in raw.items()}


def save_contractor_emails() -> None:
    if REMOTE_STORAGE:
        kv_store.set_json("contractor_emails", {str(cc): email for cc, email in state.contractor_emails.items()})
        return
    with open(CONTRACTOR_EMAILS_PATH, "w", encoding="utf-8") as f:
        json.dump({str(cc): email for cc, email in state.contractor_emails.items()}, f, indent=2, sort_keys=True)


if not REMOTE_STORAGE:
    load_contractor_emails()

ALL_10_CATEGORIES = {
    "master_data", "abnormal_shortage", "fake_documents", "quality_fail",
    "seal_tempering", "accident", "hse", "omc", "ats_trained", "reversal_cases",
}

PUBLIC_API_PATHS = {"/api/login"}

app = FastAPI(title="Cartage Contractor Performance Evaluation")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def auth_gate(request: Request, call_next):
    """Every /api/* route requires a valid bearer token from /api/login,
    except /api/login itself. The static frontend is served unauthenticated
    (it's just HTML/JS) — it can't call any API route without a token.

    In REMOTE_STORAGE mode this is also where the `state` singleton gets
    round-tripped through Vercel KV, since nothing else guarantees the same
    process (or memory) handles the next request."""
    is_api = request.url.path.startswith("/api/")
    is_public = request.url.path in PUBLIC_API_PATHS
    if is_api and not is_public:
        token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
        if not verify_token(token):
            return JSONResponse(status_code=401, content={"detail": "Not authenticated"})
        if REMOTE_STORAGE:
            state.load_dict(kv_store.get_json("app_state", {}))
            load_contractor_emails()

    response = await call_next(request)

    if REMOTE_STORAGE and is_api and not is_public:
        kv_store.set_json("app_state", state.to_dict(include_emails=False))

    return response


def _read_upload(file: UploadFile) -> io.BytesIO:
    """In-memory stand-in for a saved file path — every ingestion function
    just forwards this to pandas/openpyxl, which accept file-like objects
    the same as a path string, so nothing downstream needs to change."""
    return io.BytesIO(file.file.read())


class LoginBody(BaseModel):
    username: str
    password: str


class PeriodBody(BaseModel):
    fy: str
    period_label: str
    period_key: str


class ManualCountRow(BaseModel):
    cc_number: int
    count: int


class ManualCountsBody(BaseModel):
    rows: list[ManualCountRow]


class LimitsBody(BaseModel):
    accident: Optional[float] = None
    seal_temp: Optional[float] = None
    fake_documents: Optional[float] = None
    other_omc: Optional[float] = None
    quality_fail: Optional[float] = None
    abnormal_shortage: Optional[float] = None
    hse: Optional[float] = None


class ContractorEmailsBody(BaseModel):
    emails: dict[int, str]


class SendReportsBody(BaseModel):
    cc_numbers: list[int]


@app.post("/api/login")
def login(body: LoginBody):
    if not verify_credentials(body.username, body.password):
        raise HTTPException(401, "Invalid username or password")
    return {"token": issue_token()}


@app.post("/api/logout")
def logout(authorization: str = Header(default="")):
    token = authorization.removeprefix("Bearer ").strip()
    revoke_token(token)
    return {"ok": True}


@app.post("/api/period")
def set_period(body: PeriodBody):
    state.current_cycle_fy = body.fy.upper().replace("-", "").replace(" ", "").strip()
    state.period_label = body.period_label
    state.period_key = body.period_key
    return {"ok": True, "current_cycle_fy": state.current_cycle_fy, "period_label": state.period_label}


@app.get("/api/upload-status")
def upload_status():
    return {"logged": sorted(state.uploaded_sources), "remaining": sorted(ALL_10_CATEGORIES - state.uploaded_sources)}


@app.post("/api/upload/master-data")
def upload_master_data(file: UploadFile = File(...)):
    buf = _read_upload(file)
    try:
        result = ingest_master_data(buf)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse master data: {e}")
    state.master_table = result.as_table()
    state.uploaded_sources.add("master_data")
    return {"ok": True, "contractors_found": len(state.master_table)}


@app.post("/api/upload/abnormal-shortage")
def upload_abnormal_shortage(file: UploadFile = File(...)):
    buf = _read_upload(file)
    try:
        counts, names = ingest_abnormal_shortage(buf, state.current_cycle_fy)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse abnormal shortage file: {e}")
    state.violation_counts["abnormal_shortage"] = counts.to_dict()
    state.merge_names(names)
    state.uploaded_sources.add("abnormal_shortage")
    return {"ok": True, "contractors_found": len(counts)}


@app.post("/api/upload/fake-documents")
def upload_fake_documents(file: UploadFile = File(...)):
    buf = _read_upload(file)
    try:
        counts, names = ingest_fake_documents(buf, state.current_cycle_fy)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse fake documents file: {e}")
    state.violation_counts["fake_documents"] = counts.to_dict()
    state.merge_names(names)
    state.uploaded_sources.add("fake_documents")
    return {"ok": True, "contractors_found": len(counts)}


@app.post("/api/upload/seal-tempering")
def upload_seal_tempering(file: UploadFile = File(...)):
    buf = _read_upload(file)
    try:
        counts, names = ingest_seal_tempering(buf)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse seal tempering file: {e}")
    state.violation_counts["seal_temp"] = counts.to_dict()
    state.merge_names(names)
    state.uploaded_sources.add("seal_tempering")
    return {"ok": True, "contractors_found": len(counts)}


@app.post("/api/upload/quality-fail")
def upload_quality_fail(
    unauthorized_mod: UploadFile = File(...),
    reversal_cases: UploadFile = File(...),
):
    unauth_buf = _read_upload(unauthorized_mod)
    reversal_buf = _read_upload(reversal_cases)
    try:
        counts, names = ingest_quality_fail(unauth_buf, reversal_buf, state.current_cycle_fy)
    except Exception as e:
        raise HTTPException(400, f"Failed to parse quality fail sources: {e}")
    state.violation_counts["quality_fail"] = counts.to_dict()
    state.merge_names(names)
    state.uploaded_sources.update({"quality_fail", "reversal_cases"})
    return {"ok": True, "contractors_found": len(counts)}


@app.post("/api/manual/{category}")
def upload_manual_counts(category: str, body: ManualCountsBody):
    if category not in MANUAL_CATEGORIES:
        raise HTTPException(400, f"Unknown manual category {category!r}, expected one of {MANUAL_CATEGORIES}")
    state.manual_counts[category] = {row.cc_number: row.count for row in body.rows}
    state.uploaded_sources.add(category if category != "omc" else "omc")
    return {"ok": True, "contractors_recorded": len(body.rows)}


@app.get("/api/contractor-emails")
def get_contractor_emails():
    return {"emails": state.contractor_emails}


@app.post("/api/contractor-emails")
def upload_contractor_emails(body: ContractorEmailsBody):
    """Persisted separately from the rest of `state` (its own KV key / JSON
    file) so the address book survives a server restart / reset — set once,
    reused every evaluation cycle. An empty-string email removes that
    contractor's entry."""
    for cc, email in body.emails.items():
        if email.strip():
            state.contractor_emails[cc] = email.strip()
        else:
            state.contractor_emails.pop(cc, None)
    save_contractor_emails()
    return {"ok": True, "count": len(state.contractor_emails)}


@app.delete("/api/contractor-emails/{cc_number}")
def delete_contractor_email(cc_number: int):
    state.contractor_emails.pop(cc_number, None)
    save_contractor_emails()
    return {"ok": True, "count": len(state.contractor_emails)}


@app.post("/api/limits")
def set_limits(body: LimitsBody):
    overrides = {k: v for k, v in body.model_dump().items() if v is not None}
    state.formula_config = state.formula_config.with_threshold_overrides(overrides)
    return {"ok": True, "thresholds": state.formula_config.thresholds}


@app.post("/api/limits/reset")
def reset_limits():
    state.formula_config = default_formula_config()
    return {"ok": True, "thresholds": state.formula_config.thresholds}


@app.post("/api/calculate")
def calculate():
    rows = state.build_contractor_rows()
    if not rows:
        raise HTTPException(400, "No contractor data ingested yet — upload at least one source first.")

    scored = score_all_contractors(rows, state.formula_config)
    state.scores = {s.cc_number: s for s in scored}

    wb = build_workbook(rows, state.scores, state.previous_scores)
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    if REMOTE_STORAGE:
        current = blob_store.put_blob(f"output/scored_{state.current_cycle_fy}.xlsx", data, XLSX_MIME)
        state.output_excel_url = current["url"]
        history = blob_store.put_blob(f"history/{state.period_key}.xlsx", data, XLSX_MIME)
        manifest = load_history_manifest()
        manifest[state.period_key] = {
            "fy": state.current_cycle_fy, "period_label": state.period_label, "url": history["url"],
        }
        save_history_manifest(manifest)
    else:
        output_path = os.path.join(OUTPUT_DIR, f"scored_{state.current_cycle_fy}.xlsx")
        with open(output_path, "wb") as f:
            f.write(data)
        state.output_excel_path = output_path

        history_path = os.path.join(HISTORY_DIR, f"{state.period_key}.xlsx")
        shutil.copy2(output_path, history_path)
        manifest = load_history_manifest()
        manifest[state.period_key] = {"fy": state.current_cycle_fy, "period_label": state.period_label}
        save_history_manifest(manifest)

    return {"ok": True, "results": _results_payload()}


def _results_payload() -> list[dict]:
    payload = []
    for row in state.contractor_rows:
        cc = row["cc_number"]
        score = state.scores[cc]
        payload.append({
            "cc_number": cc,
            "name": row["name"],
            "fleet": row.get("fleet", 0),
            "overall_score": round(score.overall_score * 100, 2),
            "flagged": score.flagged,
            "category_scores": {k: round(v * 100, 2) for k, v in score.category_scores.items()},
            "category_counts": score.category_counts,
        })
    payload.sort(key=lambda r: (not r["flagged"], -r["overall_score"]))
    return payload


@app.get("/api/results")
def get_results():
    if not state.scores:
        raise HTTPException(400, "No results yet — call /api/calculate first.")
    return {"results": _results_payload()}


@app.get("/api/historical")
def get_historical():
    """Reads every archived per-period workbook back (from Vercel Blob or
    data/output/history, depending on storage mode) and returns per-
    contractor, per-category scores so the Historical Performance screen can
    chart trends across periods/browsers/sessions."""
    manifest = load_history_manifest()
    periods = []
    for period_key, meta in manifest.items():
        if REMOTE_STORAGE:
            if not meta.get("url"):
                continue
            data = blob_store.get_blob_bytes(meta["url"])
        else:
            path = os.path.join(HISTORY_DIR, f"{period_key}.xlsx")
            if not os.path.exists(path):
                continue
            with open(path, "rb") as f:
                data = f.read()

        periods.append({
            "period_key": period_key,
            "fy": meta.get("fy"),
            "period_label": meta.get("period_label"),
            "results": _parse_history_workbook(data),
        })

    return {"periods": periods}


@app.get("/api/download/excel")
def download_excel():
    if REMOTE_STORAGE:
        if not state.output_excel_url:
            raise HTTPException(400, "No scored workbook yet — call /api/calculate first.")
        data = blob_store.get_blob_bytes(state.output_excel_url)
        filename = f"scored_{state.current_cycle_fy}.xlsx"
        return Response(
            content=data, media_type=XLSX_MIME,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if not state.output_excel_path or not os.path.exists(state.output_excel_path):
        raise HTTPException(400, "No scored workbook yet — call /api/calculate first.")
    return FileResponse(
        state.output_excel_path,
        media_type=XLSX_MIME,
        filename=os.path.basename(state.output_excel_path),
    )


@app.get("/api/reports/preview/{cc_number}")
def preview_report(cc_number: int):
    if cc_number not in state.scores:
        raise HTTPException(404, "Unknown contractor — run /api/calculate first.")
    score = state.scores[cc_number]
    name = state.names.get(cc_number, str(cc_number))
    data = generate_report(cc_number, name, score, state.period_label)
    filename = f"violation_report_{cc_number}.docx"
    return Response(
        content=data, media_type=DOCX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/reports/send")
def send_reports(body: SendReportsBody):
    """Only sends to contractors explicitly approved in the request body — the
    frontend's review queue (Section 10/12: no auto-send). Each cc_number is
    re-checked against the 80% threshold before sending regardless of what the
    client requests. Reports are regenerated fresh here rather than reused
    from a previous request — cheap to build, and nothing to keep alive
    across serverless invocations that way."""
    results = []
    for cc in body.cc_numbers:
        if cc not in state.scores:
            results.append({"cc_number": cc, "success": False, "error": "unknown contractor"})
            continue
        score = state.scores[cc]
        if not score.flagged:
            results.append({"cc_number": cc, "success": False, "error": "overall_score >= 80%, not eligible"})
            continue

        name = state.names.get(cc, str(cc))
        report_bytes = generate_report(cc, name, score, state.period_label)

        result = send_report_email(
            cc_number=cc,
            name=name,
            overall_score=score.overall_score,
            contractor_email=state.contractor_emails.get(cc),
            report_bytes=report_bytes,
            report_name=f"violation_report_{cc}.docx",
            period=state.period_label,
        )
        results.append({
            "cc_number": cc, "email": result.email, "success": result.success, "error": result.error,
        })

    sent = sum(1 for r in results if r["success"])
    return {"sent": sent, "failed": len(results) - sent, "details": results}


@app.post("/api/reset")
def reset_state():
    state.reset()
    load_contractor_emails()  # the address book persists across resets, unlike the run data
    return {"ok": True}


# Serves the frontend (index.html + assets) from the same origin as the API,
# so there is nothing to configure when moving this to another machine — one
# process, one port, no CORS/base-URL setup. Mounted last so it never shadows
# an /api/* route above.
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
if os.path.isdir(FRONTEND_DIR):
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
