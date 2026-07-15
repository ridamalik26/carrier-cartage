"""Fake / Invalid Driving License ingestion (Section 5.2).

Sheet name has a trailing space in the real file ("Datewise Data ") and the
header sits on row 3 (rows 1-2 are a title + spacer row), so the header row
is auto-detected by scanning for "Sr. No." rather than hardcoding an offset.
"""
from __future__ import annotations

import pandas as pd
import openpyxl

from backend.ingestion.generic import ViolationSource, ingest_violation_count


def _find_sheet(file_path: str, expected: str = "Datewise Data") -> str:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    try:
        for sheet in wb.sheetnames:
            if sheet.strip() == expected:
                return sheet
        raise ValueError(f"No sheet matching {expected!r} (stripped) found among {wb.sheetnames}")
    finally:
        wb.close()


def ingest_fake_documents(file_path: str, current_cycle_fy: str | None = None) -> tuple[pd.Series, dict[int, str]]:
    sheet_name = _find_sheet(file_path)
    source = ViolationSource(
        name="fake_documents",
        file_path=file_path,
        sheet_name=sheet_name,
        id_column="Carriage Code",
        header_row=None,
        header_anchor="Sr. No.",
        fy_column="FY",
        current_cycle_fy=current_cycle_fy,
    )
    return ingest_violation_count(source)
